"""
Extractor de XLSX usando openpyxl.

Extrae:
- Múltiples hojas (sheets)
- Tablas con headers normalizados
- Merged cells detectadas
- Metadata del documento
"""

import logging
from typing import List, Dict, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

_logger = logging.getLogger(__name__)


class XLSXExtractor:
    """
    Extractor especializado para archivos XLSX.
    
    Usa openpyxl para:
    - Extracción de múltiples hojas
    - Detección de merged cells
    - Normalización de headers
    - Metadata del documento
    """
    
    def __init__(self):
        """Inicializa el extractor XLSX."""
        _logger.info("📊 XLSXExtractor inicializado")
    
    def extract(self, xlsx_path: str) -> Dict:
        """
        Extrae todo el contenido de un XLSX.
        
        Args:
            xlsx_path: Ruta local del archivo XLSX
        
        Returns:
            Diccionario con:
            {
                "metadata": {...},
                "sheets": [
                    {
                        "name": "Sheet1",
                        "headers": [...],
                        "rows": [[...]],
                        "merged_cells": [...],
                        "num_rows": 100,
                        "num_cols": 10
                    },
                    ...
                ],
                "extraction_method": "openpyxl"
            }
        
        Raises:
            ValueError: Si el archivo no es válido
            FileNotFoundError: Si el archivo no existe
        """
        try:
            _logger.info(f"📖 Extrayendo XLSX: {xlsx_path}")
            
            if not Path(xlsx_path).exists():
                raise FileNotFoundError(f"Archivo no encontrado: {xlsx_path}")
            
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            
            # Extraer metadata
            metadata = self._extract_metadata(wb)
            
            # Extraer todas las hojas
            sheets = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = self._extract_sheet(sheet)
                sheets.append(sheet_data)
            
            result = {
                "metadata": metadata,
                "sheets": sheets,
                "total_sheets": len(wb.sheetnames),
                "extraction_method": "openpyxl"
            }
            
            _logger.info(
                f"✅ XLSX extraído: {len(sheets)} hojas"
            )
            
            return result
        
        except Exception as e:
            _logger.error(f"❌ Error extrayendo XLSX: {e}", exc_info=True)
            raise
    
    def _extract_metadata(self, wb) -> Dict:
        """
        Extrae metadata del workbook.
        
        Args:
            wb: Objeto Workbook de openpyxl
        
        Returns:
            Metadata del documento
        """
        try:
            props = wb.properties
            return {
                "title": props.title or "",
                "creator": props.creator or "",
                "created": str(props.created) if props.created else "",
                "modified": str(props.modified) if props.modified else "",
                "sheet_names": wb.sheetnames
            }
        except Exception as e:
            _logger.warning(f"⚠️ Error extrayendo metadata: {e}")
            return {"sheet_names": wb.sheetnames}
    
    def _extract_sheet(self, sheet) -> Dict:
        """
        Extrae contenido de una hoja.
        
        Args:
            sheet: Objeto Worksheet de openpyxl
        
        Returns:
            Datos de la hoja con headers, rows, merged cells
        """
        try:
            sheet_name = sheet.title
            _logger.debug(f"📄 Procesando hoja: {sheet_name}")
            
            # Detectar merged cells
            merged_cells = self._detect_merged_cells(sheet)
            
            # Extraer datos como lista de listas
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Convertir None a string vacío
                row_data = [str(cell) if cell is not None else "" for cell in row]
                data.append(row_data)
            
            # Determinar headers (primera fila no vacía)
            headers = []
            rows = []
            
            if data:
                # Buscar primera fila no vacía como headers
                for i, row in enumerate(data):
                    if any(cell.strip() for cell in row):
                        headers = row
                        rows = data[i+1:]
                        break
            
            # Filtrar filas completamente vacías
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            
            return {
                "name": sheet_name,
                "headers": headers,
                "rows": rows,
                "merged_cells": merged_cells,
                "num_rows": len(rows),
                "num_cols": len(headers) if headers else 0
            }
        
        except Exception as e:
            _logger.error(f"❌ Error extrayendo hoja {sheet.title}: {e}")
            return {
                "name": sheet.title,
                "headers": [],
                "rows": [],
                "merged_cells": [],
                "num_rows": 0,
                "num_cols": 0,
                "error": str(e)
            }
    
    def _detect_merged_cells(self, sheet) -> List[Dict]:
        """
        Detecta celdas combinadas (merged cells).
        
        Args:
            sheet: Objeto Worksheet
        
        Returns:
            Lista de rangos merged: [{"range": "A1:B2", "value": "..."}]
        """
        merged = []
        
        try:
            for merged_range in sheet.merged_cells.ranges:
                # Obtener valor de la primera celda del rango
                min_col, min_row, max_col, max_row = merged_range.bounds
                first_cell = sheet.cell(min_row, min_col)
                
                merged.append({
                    "range": str(merged_range),
                    "value": str(first_cell.value) if first_cell.value else "",
                    "rows": max_row - min_row + 1,
                    "cols": max_col - min_col + 1
                })
        
        except Exception as e:
            _logger.warning(f"⚠️ Error detectando merged cells: {e}")
        
        return merged
    
    def extract_sheet_as_table(self, xlsx_path: str, sheet_name: Optional[str] = None) -> Dict:
        """
        Extrae una hoja específica como tabla.
        
        Args:
            xlsx_path: Ruta del XLSX
            sheet_name: Nombre de la hoja (None = primera hoja)
        
        Returns:
            Tabla con headers y rows
        """
        full_extraction = self.extract(xlsx_path)
        
        if not full_extraction["sheets"]:
            return {"headers": [], "rows": []}
        
        # Si no se especifica, tomar primera hoja
        if sheet_name is None:
            return full_extraction["sheets"][0]
        
        # Buscar hoja específica
        for sheet in full_extraction["sheets"]:
            if sheet["name"] == sheet_name:
                return sheet
        
        raise ValueError(f"Hoja '{sheet_name}' no encontrada")
    
    def extract_all_text(self, xlsx_path: str) -> str:
        """
        Extrae todo el texto plano de todas las hojas.
        
        Args:
            xlsx_path: Ruta del XLSX
        
        Returns:
            Texto plano concatenado
        """
        full_extraction = self.extract(xlsx_path)
        
        text_parts = []
        for sheet in full_extraction["sheets"]:
            text_parts.append(f"=== {sheet['name']} ===")
            
            # Headers
            if sheet["headers"]:
                text_parts.append(" | ".join(sheet["headers"]))
            
            # Rows (primeras 100 para no exceder)
            for row in sheet["rows"][:100]:
                text_parts.append(" | ".join(row))
        
        return "\n".join(text_parts)


if __name__ == "__main__":
    # Test básico
    import sys
    logging.basicConfig(level=logging.INFO)
    
    if len(sys.argv) > 1:
        extractor = XLSXExtractor()
        result = extractor.extract(sys.argv[1])
        print(f"\n✅ Extraído: {result['total_sheets']} hojas")
        
        for sheet in result['sheets']:
            print(f"\n📄 Hoja: {sheet['name']}")
            print(f"   Dimensiones: {sheet['num_rows']} filas x {sheet['num_cols']} columnas")
            print(f"   Merged cells: {len(sheet['merged_cells'])}")
            if sheet['headers']:
                print(f"   Headers: {', '.join(sheet['headers'][:5])}...")
    else:
        print("Uso: python xlsx_extractor.py <ruta_xlsx>")

